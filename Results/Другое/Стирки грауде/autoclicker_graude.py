import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === настройки ===
excel_file = "articles.xlsx"   # твой Excel с артикулами
txt_file   = "links_graude.txt"       # куда пишем ссылки
site_url   = "https://graude-shop.ru/"  # базовый адрес

# цвета для Excel
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# === браузер ===
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # если нужно без окна
driver = webdriver.Chrome(options=options)

# === открываем Excel ===
wb = load_workbook(excel_file)
ws = wb.active

links_out = []
deffects_out = []

# === обходим строки с артикулами ===
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=2), start=2):
    cell = row[0]
    article = str(cell.value).strip() if cell.value else ""
    link = ""
    
    if not article:
        links_out.append("")  # пустая строка
        continue

    try:
        # Переход на сайт
        driver.get(site_url)
        
        # Используем execute_script для всех взаимодействий
        # Очистка поля поиска и ввод текста
        driver.execute_script("""
            document.querySelector(".btn.btn--no-rippple.btn-clear-search").click();
            var inputField = document.querySelector('input.form-control.search-input');
            inputField.value = arguments[0];
            inputField.dispatchEvent(new Event('change'));
            inputField.dispatchEvent(new KeyboardEvent('keydown', {'keyCode': 13}));
        """, article)
        
        # Ждем появления карточек товаров
        result_cards = driver.execute_script("""
            return Array.from(document.querySelectorAll('div.image-list-wrapper.js-image-block')).length;
        """)
        
        if result_cards != 1:
            deffects_out.append(f"{article} - {idx}")
            cell.fill = fill_red
            links_out.append("")
        else:
            # Получаем первую карточку товара и извлекаем её ссылку
            link = driver.execute_script("""
                const card = document.querySelector('div.image-list-wrapper.js-image-block a');
                return card ? card.href : '';
            """)
            
            if link:
                cell.fill = fill_yellow
                links_out.append(link)
            else:
                deffects_out.append(f"{article} - {idx}")
                cell.fill = fill_red
                links_out.append("")
                
    except Exception as e:
        print(f"Ошибка для артикула {article}: {e}")
        cell.fill = fill_red
        links_out.append("")

# === сохраняем Excel ===
wb.save(excel_file)

# === пишем txt ===
with open(txt_file, "w", encoding="utf-8") as f:
    for l in links_out:
        f.write(l + "\n")

# === пишем deffects.txt ===
with open("deffects_graude.txt", "w", encoding="utf-8") as f:
    for d in deffects_out:
        f.write(d + "\n")

driver.quit()
