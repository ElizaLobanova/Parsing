import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === настройки ===
excel_file = "articles.xlsx"   # твой Excel с артикулами
txt_file   = "links_housedorf.txt"       # куда пишем ссылки
site_url   = "https://www.hausdorf.ru/"  # базовый адрес

# цвета для Excel
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# === браузер ===
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # если нужно без окна
driver = webdriver.Chrome(options=options)

# === открываем Excel ===
wb = load_workbook(excel_file)
ws = wb.active

links_out = []
deffects_out = []

# === обходим строки с артикулами ===
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=2), start=2):  # допустим, артикулы в 1-й колонке
    cell = row[0]
    article = str(cell.value).strip() if cell.value else ""
    link = ""

    if not article:
        links_out.append("")  # пустая строка
        continue

    try:
        # заходим на сайт
        driver.get(site_url)
        time.sleep(2)

        # ищем строку поиска
        search_box = driver.find_element(By.NAME, "q")  # нужно уточнить имя поля
        # search_box.clear()
        search_box.send_keys(article)
        search_box.send_keys(Keys.RETURN)
        time.sleep(2)

        # ищем карточки в выдаче
        swipers = driver.find_elements(By.CSS_SELECTOR, "div.swiper")

        if len(swipers) != 1:
            # пишем в defеcts.txt
            deffects_out.append(f"{article} - {idx}")
            cell.fill = fill_red
            links_out.append("")
        else:
            # берём ссылку
            try:
                link = swipers[0].find_element(By.TAG_NAME, "a").get_attribute("href")
                cell.fill = fill_yellow
                links_out.append(link)
            except:
                deffects_out.append(f"{article} - {idx}")
                cell.fill = fill_red
                links_out.append("")

    except Exception as e:
        print(f"Ошибка для артикула {article}: {e}")
        cell.fill = fill_red
        links_out.append("")

# === сохраняем Excel ===
wb.save(excel_file)

# === пишем txt ===
with open(txt_file, "w", encoding="utf-8") as f:
    for l in links_out:
        f.write(l + "\n")

# === пишем deffects.txt ===
with open("deffects_housedorf.txt", "w", encoding="utf-8") as f:
    for d in deffects_out:
        f.write(d + "\n")

driver.quit()
