import urllib3
from bs4 import BeautifulSoup, NavigableString
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
import re
from tqdm import tqdm
import numpy as np
import warnings
warnings.filterwarnings("ignore")
import argparse
import os
from ruwordnet import RuWordNet
import pymorphy2
from itertools import product
from pathlib import Path

synonyms_path='synonyms.txt'
all_characteristics_path='all_characteristics.xlsx'

# ---------------------------------------------------Остановить программу если не введено ни одного аргумента---------------------------------------------------
parser = argparse.ArgumentParser(description="Парсинг данных в excel-файл с одного типа сайтов. В качестве входного файла используется выгрузка из 1С. " \
"Выходной файл создаётся по образу и подобию входного, является результатом парсинга. Если какая-то ячейка уже была заполнена в excel-файле, то она не " \
"будет перезаписана. Возможно дополнение файла столбцами на основе найденных на сайте характеристик. Дополнительно сохраняется вспомогательный .parquet-файл " \
"для дальнейшего сравнения данных с разных сайтов, а также .txt-файл для дальнейшей генерации отчёта о синонимах.")

parser.add_argument("start_row", type=int, help="Номер строки в excel, начиная с которой необходимо писать данные")
parser.add_argument("append", type=str, help="Добавлять ли в конец дополнительные столбцы с незаписанными данными сайтов. Возможные значения: True, " \
"False. Если False, то незаписанные данные будут сохранены в отдельный excel-файл с названием, начинающимся с 'missing'. Название характреристик в этом файле " \
"будут приведены к синонимичным из 1С в соответствии с утверждённым словарём синонимов. ")
parser.add_argument("site", type=str, help='Название типа сайта для парсинга. Возможные значения: korting, housedorf, dedietrich, falmec, vzug, asco, kuppersbush, konigin, evelux. ')
parser.add_argument("urls_source", type=str, help='Файл со ссылками на карточки товаров. Порядок должен соответствовать расположению наименований ' \
'номенклатуры в excel-файле, указанном как входной')
parser.add_argument("input_path", type=str, help='Путь к входному excel-файлу')
parser.add_argument("output_path", type=str, help='Путь к выходному excel-файлу')

args = parser.parse_args()
if not args.output_path:
    raise ValueError("there's not enough arguments")
args.append = True if args.append.lower() == 'true' else False

# Проверка, не повреждены ли входные файлы
wb = load_workbook(args.input_path)

# ------------------------------------------------------------------------Спарсить данные------------------------------------------------------------------------

# Функции для парсинга
def parse_korting_page(html_code):
    soup = BeautifulSoup(html_code, 'html.parser')
    tabs_lists = soup.find_all('ul', class_='tabs-settings__list')
    data = {}

    for ul in tabs_lists:
        for li in ul.find_all('li'):
            text = li.get_text(strip=True, separator="; ")
            split_text = text.split(":;", 1)
            if len(split_text) == 2:
                key, value = split_text
                data[key.strip()] = value.strip()
            else:
                data[split_text[0].strip()] = ""

    return data

def parse_dedietrich_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # Проходим по всем div с классом characteristics__row
    for row in soup.find_all('div', class_='characteristics__row'):
        name_span = row.find('span', class_='characteristics__name')
        value_span = row.find('span', class_='characteristics__property')
        
        if name_span and value_span:
            key = name_span.find(text=True, recursive=False)
            value = value_span.find(text=True, recursive=False)
            if key and value:
                key = key.strip()
                value = value.strip()
                data[key] = value

    return data

def parse_vzug_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}
    names = soup.find_all('td', class_='cell_name')
    values = soup.find_all('td', class_='cell_value')

    # Проходим по всем div с классом characteristics__row
    for name, value in zip(names, values):
        name_span = name.find('span')
        value_span = value.find('span')
        
        if name_span and value_span:
            key = name_span.find(text=True, recursive=False)
            value = value_span.find(text=True, recursive=False)
            if key and value:
                key = key.strip()
                value = value.strip()
                data[key] = value

    return data

def parse_asco_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # Проходим по всем div с классом characteristics__row
    for row in soup.find_all('div', class_='accordeon__item'):
        name_span = row.find('span', class_='accordeon__item-title')
        value_span = row.find('p', class_='accordeon__item-text')
        
        if name_span and value_span:
            key = name_span.find(text=True, recursive=False)
            value = value_span.find(text=True, recursive=False)
            if key and value:
                key = key.strip()
                value = value.strip()
                data[key] = value

    return data

def parse_kuppersbush_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # Проходим по всем div с классом characteristics__row
    for row in soup.find_all('div', class_='wdu_propsorter'):
        for item in row.find_all('tr'):
            tds = item.find_all('td')

            if tds:
                name_td, value_td = tds
                key = name_td.get_text(strip=True)
                value = value_td.get_text(strip=True)
                if key and value:
                    data[key] = value

    return data

def parse_evelux_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # Проходим по всем div с классом characteristics__row
    for row in soup.find_all('div', class_='product__content-specs-line'):
        name_span = row.find('div', class_='product__content-specs-title')
        value_span = row.find('div', class_='product__content-specs-subtitle')
        
        if name_span and value_span:
            key = name_span.find(text=True, recursive=False).split(':')[0]
            value = value_span.find(text=True, recursive=False)
            if key and value:
                key = key.strip()
                value = value.strip()
                data[key] = value

    return data

def parse_konigin_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # 1. Извлекаем div.column и все p внутри
    column_div = soup.find('div', class_='column')
    if column_div:
        p_tags = column_div.find_all('p')
        # Игнорируем первые 2 и последние 3
        middle_p_tags = p_tags[2:-3]

        # Соединяем текст через "; "
        additional_info = "; ".join(p.get_text(strip=True).rstrip('.,;!?—:') for p in middle_p_tags)
        data['ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ'] = additional_info

        # Обрабатываем последние два тега <p>
        for p in p_tags[-2:]:
            text = p.get_text(strip=True)
            # Делим по точкам
            for part in text.split('.'):
                if ':' in part:
                    key, value = part.split(':', 1)
                    data[key.strip()] = value.strip()

    # 2. Обрабатываем блоки div.card-spec__item
    spec_items = soup.find_all('div', class_='card-spec__item')
    for item in spec_items:
        title_div = item.find('div', class_='card-spec__item-title')
        info_div = item.find('div', class_='card-spec__item-info')
        if title_div and info_div:
            key = title_div.get_text(strip=True)
            value = info_div.get_text(strip=True).rstrip('.,;!?—:')
            data[key] = value

    return data

def parse_falmec_page(html_code: str) -> dict:
    soup = BeautifulSoup(html_code, 'html.parser')
    data = {}

    # Проходим по всем div с классом characteristics__row
    for row in soup.find_all('div', class_='characteristics__row'):
        name_span = row.find('span', class_='characteristics__name')
        value_span = row.find('span', class_='characteristics__property')
        
        if name_span and value_span:
            key = name_span.find(text=True, recursive=False)
            value = value_span.find(text=True, recursive=False)

            # Если нет простого текста, ищем <ul> и собираем <li>
            if (value == '' or not value.strip()) and value_span.find('ul'):
                li_items = value_span.find_all('li')
                value = '; '.join(li.get_text(strip=True) for li in li_items)

            if key and value:
                key = key.strip()
                value = value.strip()
                data[key] = value

    return data

def extract_first_visible_text(tag):
    for desc in tag.descendants:
        if isinstance(desc, str):  # Это NavigableString
            text = desc.strip()
            if text:
                return text
    return None

def clean_value_div(value_div):
    # 1. Удалить все <span>
    for span in value_div.find_all("span"):
        span.decompose()

    # 2. Разделить по <br> — создаём список на основе HTML с разделителем
    parts = str(value_div).split('<br')

    values = []

    for part_html in parts:
        # Восстанавливаем HTML-тег <br>, если он был отрезан
        if not part_html.startswith('>'):
            part_html = '<br' + part_html

        part_soup = BeautifulSoup(part_html, 'html.parser')

        # 3. Найти первый видимый текст
        for desc in part_soup.descendants:
            if isinstance(desc, NavigableString):
                text = desc.strip()
                if text:
                    values.append(text)
                    break  # только первое вхождение

    # 4. Склеить с разделителем "; "
    return "; ".join(values)

def parse_hausedorf_page(html_code):
    soup = BeautifulSoup(html_code, 'html.parser')
    fields = soup.find_all('div', class_='detail-properties__field')
    data = {}

    for field in fields:
        name_div = field.find('div', class_='detail-properties__name')
        value_div = field.find('div', class_='detail-properties__value')

        if name_div and value_div:
            key = extract_first_visible_text(name_div)
            value = clean_value_div(value_div)

            if key and value:
                data[re.sub(r'\s+', ' ', key).strip()] = value.replace(">\n", "")

    return data


def create_src(file_path, parser_func):
    """
    Универсальный загрузчик таблицы характеристик с разных сайтов.

    :param file_path: путь к файлу с URL (один URL на строку)
    :param parser_func: функция, которая получает HTML-код и возвращает словарь {ключ: значение}
    :return: DataFrame с объединёнными результатами
    """
    http = urllib3.PoolManager()
    df_all = pd.DataFrame()

    with open(file_path, 'r', encoding='utf-8') as f:
        urls = [line.strip() for line in f]

    n_rows = 0
    for url in tqdm(urls):
        if url:
            try:
                response = http.request('GET', url)
                html_code = response.data.decode()
                data = parser_func(html_code) 

                if not isinstance(data, dict):
                    raise ValueError("parser_func должна возвращать словарь!")

                row_df = pd.DataFrame([data])
                df_all = pd.concat([df_all, row_df], ignore_index=True)

                if len(df_all) == 1:
                    empty_rows = pd.DataFrame(np.nan, index=range(n_rows), columns=df_all.columns)
                    df_all = pd.concat([empty_rows, df_all], ignore_index=True)

            except Exception as e:
                print(f"Ошибка при обработке {url}: {e}")
        else:
            if len(df_all.columns) > 0:                
                df_all.loc[len(df_all)] = None
            else:
                n_rows += 1

    return df_all.where(pd.notnull(df_all), None)

# ---------------------------------------------Записать и вернуть дополненный названиями номенклатуры DataFrame---------------------------------------------

def write_dest(ref_file_path, result_file_path, df_src, start_row_index):
    # Путь к файлу Excel
    wb = load_workbook(ref_file_path)
    ws = wb.active  # или wb['SheetName']

    # Поля файла-приёмника
    row_header = [cell.value for cell in ws[1]]

    # Сопоставление колонок
    src_cols_lower = {col.lower(): col for col in df_src.columns}
    ws_cols_lower = {i: str(header).strip().lower() if header else "" for i, header in enumerate(row_header)}
    matched_columns = []
    common_cols = set()
    nomenclature_col_idx = None

    for col_idx, header_lower in ws_cols_lower.items():
        if header_lower == "номенклатура":
            nomenclature_col_idx = col_idx
        if header_lower in src_cols_lower:
            matched_columns.append((col_idx, src_cols_lower[header_lower]))
            common_cols.add(src_cols_lower[header_lower])

    if nomenclature_col_idx is None:
        raise ValueError("Колонка 'Номенклатура' не найдена в файле-приёмнике")

    # Считываем значения "Номенклатура"
    nomenclature_values = []
    for i in range(len(df_src)):
        cell_value = ws.cell(row=start_row_index + i, column=nomenclature_col_idx + 1).value
        nomenclature_values.append(cell_value)

    # Запись данных
    for i, (_, row_src) in enumerate(df_src.iterrows()):
        for col_idx, src_col in matched_columns:
            cell = ws.cell(row=start_row_index + i, column=col_idx + 1)
            if cell.value in [None, ""]:
                cell.value = row_src[src_col]

    # Сохранение
    wb.save(result_file_path)

    # Подготовка выходного DataFrame
    result_df = df_src[list(common_cols)].copy()
    result_df.insert(0, "Номенклатура", pd.Series(nomenclature_values))

    return result_df, common_cols

# -------------------------------------------Сохранить незаписанные данные в дополнительные колонки или отдельный файл-------------------------------------------

def append_dataframe_to_excel(df: pd.DataFrame, file_path: str, result_path: str, start_row: int):
    # Проверка, существует ли файл
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        # Если файла нет, создаем новый
        wb = Workbook()
    
    ws = wb.active

    # Найдём первую пустую ячейку в первой строке
    col_index = 1
    while ws.cell(row=1, column=col_index).value is not None:
        col_index += 1

    # Записываем названия колонок DataFrame в первую строку, начиная с найденной колонки
    for i, col_name in enumerate(df.columns):
        cell = ws.cell(row=1, column=col_index + i, value=col_name)
        if len(col_name.split('_')) > 1:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            

    # Автонастройка ширины столбцов по первой строке
    for col_idx, cell in enumerate(ws[1], start=col_index):
        max_length = len(str(cell.value)) if cell.value else 0
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = max_length + 2  # +2 для отступа

    # Применение стилей и переносов
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')  # включаем перенос текста

    # Записываем данные DataFrame начиная со start_row
    for row_offset, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for i, value in enumerate(row):
            ws.cell(row=start_row + row_offset, column=col_index + i, value=value)

    # Сохраняем файл
    wb.save(result_path)

def save_missing(df1, filepath):

    # Создаём ExcelWriter
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        row = 0

        # Запись df1
        df1.to_excel(writer, index=False, startrow=row)
        row += len(df1) + 2  # +1 за заголовок, +1 за пустую строку

        # Автоматическая установка ширины колонок
        worksheet = writer.sheets['Sheet1']
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

# ---------------------------------------------------------Добавить функции для сопоставления синонимов---------------------------------------------------------

wordnet = RuWordNet()
morph = pymorphy2.MorphAnalyzer()

def get_normal_form(word):
    return morph.parse(word)[0].normal_form

def are_synonyms(word1, word2):
    lemma1 = get_normal_form(word1)
    lemma2 = get_normal_form(word2)

    synsets1 = wordnet.get_synsets(lemma1)
    synsets2 = wordnet.get_synsets(lemma2)

    # Сравниваем наличие общих лемм в синсетах
    for s1 in synsets1:
        for s2 in synsets2:
            if s1.id == s2.id:
                return True
    return False

def list_synonyms_comparison(list1, list2):
    return [are_synonyms(word1, word2) for word1, word2 in zip(list1, list2)]

# -----------------------------Добавить функцию для обеспечения правильного дописывания данных основываясь на утверждённых синонимах-----------------------------

def parse_custom_dict_line(line):
    """
    Разбирает строку из словаря: <характеристика>: <синоним1>; <синоним2>; ...| <антисиноним1>, <антисиноним2>, ...
    """
    base, *rest = line.strip().split(':')
    if not rest:
        return base.strip(), set(), set()
    syn_ant = rest[0].split('|')
    synonyms = set(map(str.strip, syn_ant[0].split(';'))) if syn_ant[0] else set()
    antisynonyms = set(map(str.strip, syn_ant[1].split(','))) if len(syn_ant) > 1 else set()
    return base.strip(), synonyms, antisynonyms

def load_existing_synonyms(file_path):
    syn_dict = {}
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            base, synonyms, antisynonyms = parse_custom_dict_line(line)
            syn_dict[base] = {'synonyms': synonyms, 'antisynonyms': antisynonyms}
    return syn_dict

def rename_columns_with_syn_dict(df, syn_dict_path, all_1c_chars_path):
    # Загрузка словаря
    synonyms_dict = load_existing_synonyms(syn_dict_path)
    all_chars = pd.read_excel(all_1c_chars_path, header=None).iloc[0].astype(str).tolist()
    all_chars_lower = [char.lower() for char in all_chars]  # Приводим к нижнему регистру для сравнения

    # Удаление дублирующихся колонок по имени
    df = df.loc[:, ~df.columns.duplicated()]

    # Поиск синонимичных имен колонок через словарь
    df_expanded = df.copy()
    unsyn_set = set()
    for col in df.columns:
        col_splitted = col.split(':')[0]
        # Если колонка уже есть в 1С - пропускаем её
        if col_splitted.lower() in all_chars_lower:
            df_expanded[col_splitted.upper()] = df[col]
            df_expanded.drop(columns=[col], inplace=True)
            continue

        is_syn = False
        for synonym_key, syn_data in synonyms_dict.items():
            syn_set = syn_data.get("synonyms", set())
            if col_splitted in syn_set:
                # Добавляем колонку с именем synonym_key, если она ещё не существует
                if synonym_key not in df_expanded.columns:
                    df_expanded[synonym_key] = df[col]
                    is_syn = True
        if is_syn:
            # Если колонка была переименована, удаляем оригинальную и не ищем полусинонимы
            df_expanded.drop(columns=[col], inplace=True)
        else:
            is_half_syn = False
            for synonym_key, syn_data in synonyms_dict.items():
                syn_set = syn_data.get("synonyms", set())
                if col_splitted in set(map(lambda x: x.split("*")[1] if len(x.split("*")) > 1 else x, syn_set)):
                    # Добавляем колонку с именем synonym_key, если она ещё не существует
                    if synonym_key not in df_expanded.columns:
                        df_expanded[f"{synonym_key}_{col_splitted}"] = df[col]
                        is_half_syn = True

            if is_half_syn:
                df_expanded.drop(columns=[col], inplace=True)
            else:
                is_antisyn = False
                for synonym_key, syn_data in synonyms_dict.items():
                    antisyn_set = syn_data.get("antisynonyms", set())
                    if col_splitted in antisyn_set:
                        is_antisyn = True
                        break

                if not is_antisyn:
                    unsyn_set.add(col_splitted)

    return df_expanded, unsyn_set

# -----------------------------------------------------------Запуск парсинга и сохранение результатов-----------------------------------------------------------

if args.site == 'korting':
    df_src = create_src(args.urls_source, parse_korting_page)
elif args.site == 'housedorf':
    df_src = create_src(args.urls_source, parse_hausedorf_page)
elif args.site == 'dedietrich':
    df_src = create_src(args.urls_source, parse_dedietrich_page)
elif args.site == 'falmec':
    df_src = create_src(args.urls_source, parse_falmec_page)
elif args.site == 'vzug':
    df_src = create_src(args.urls_source, parse_vzug_page)
elif args.site == 'asco':
    df_src = create_src(args.urls_source, parse_asco_page)
elif args.site == 'kuppersbush':
    df_src = create_src(args.urls_source, parse_kuppersbush_page)
elif args.site == 'konigin':
    df_src = create_src(args.urls_source, parse_konigin_page)
elif args.site == 'evelux':
    df_src = create_src(args.urls_source, parse_evelux_page)
else:
    raise ValueError("There're no parse function for this site")

df_src, unsyn_set = rename_columns_with_syn_dict(df_src, synonyms_path, all_characteristics_path) # Для обеспечения правильного дописывания данных и для корректного входа к функции, 
                                                                        # генерирующей отчёт
resultdf, com_cols = write_dest(args.input_path, args.output_path, df_src, args.start_row)
resultdf.to_parquet(f"{args.site}_auxiliary.parquet")

com_cols = resultdf.columns.intersection(df_src.columns)
missingdf = df_src.drop(columns=com_cols).copy()
print(unsyn_set)
with open(f'unaccepted_syn_{args.site}.txt', 'w', encoding='utf-8') as f:
    f.write('; '.join(map(str, unsyn_set)))

if args.append:
    append_dataframe_to_excel(missingdf, args.output_path, args.output_path, args.start_row)
else:
    # Условие: имена столбцов, у которых есть хотя бы один символ "_"
    columns_with_underscore = [col for col in missingdf.columns if len(col.split("_")) > 1]
    columns_without_underscore = [col for col in missingdf.columns if len(col.split("_")) <= 1]

    # Делим missingdf на два
    df_with_underscore = missingdf[columns_with_underscore]
    df_without_underscore = missingdf[columns_without_underscore]
    df_without_underscore.insert(0, 'Номенклатура', resultdf['Номенклатура'].copy())
    save_missing(df_without_underscore, f'missing_{args.site}.xlsx')
    append_dataframe_to_excel(df_with_underscore, args.output_path, args.output_path, args.start_row)
    
print("Successfully finished")