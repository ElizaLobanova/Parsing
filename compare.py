import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse

# ---------------------------------------------------Остановить программу если не введено ни одного аргумента---------------------------------------------------
# Номер строки, взятый из аргументов запуска программы
parser = argparse.ArgumentParser(description="Сравнивает результаты парсинга с двух сайтов (убедитесь, что парсинг с каждого из указываемых в аргументах " \
"к этой программе сайтов проведён с помощью программы parsing.py). Программа использует сохранённые программой парсинга вспомогательные файлы. " \
"Результат сравнения записывается в excel-файл с названием, начинающемся с 'comparison', с упоминанием сайтов, по которым шло сравнение.")

parser.add_argument("site1", type=str, help="Тип первого сайта, участвующего в сравнении")
parser.add_argument("site2", type=str, help='Тип второго сайта, участвующего в сравнении')

args = parser.parse_args()
if not args.site2:
    raise ValueError("there's not enough arguments")
if args.site1 == args.site2:
    raise ValueError("the arguments coincide, comparison is impossible")

# -------------------------------------------------------------------Сравнить записанные данные-------------------------------------------------------------------

def compare_dataframes(df1: pd.DataFrame, df2: pd.DataFrame, name1: str = 'df1', name2: str = 'df2') -> pd.DataFrame:
    """
    Сравнивает два DataFrame по общим столбцам, включая "Номенклатура" как ключ.
    Добавляет префиксы к столбцам (кроме "Номенклатура") и формирует колонку diff_columns.
    """
    if 'Номенклатура' not in df1.columns or 'Номенклатура' not in df2.columns:
        raise ValueError("Оба DataFrame должны содержать колонку 'Номенклатура'")

    # Определим общие характеристики (кроме "Номенклатура")
    common_columns = df1.columns.intersection(df2.columns).difference(['Номенклатура'])

    # Сузим входные DataFrame'ы до нужных колонок
    df1_reduced = df1[['Номенклатура'] + list(common_columns)].copy()
    df2_reduced = df2[['Номенклатура'] + list(common_columns)].copy()

    # Переименуем характеристики с префиксами, "Номенклатура" оставим без изменений
    df1_renamed = df1_reduced.rename(columns={col: f'{name1}_{col}' for col in common_columns})
    df2_renamed = df2_reduced.rename(columns={col: f'{name2}_{col}' for col in common_columns})

    # Объединение по "Номенклатура"
    df_merged = pd.merge(df1_renamed, df2_renamed, on='Номенклатура', how='outer')

    # Функция для сравнения значений по строке
    def get_differences(row):
        diffs = []
        for col in common_columns:
            val1 = row.get(f'{name1}_{col}', None)
            val2 = row.get(f'{name2}_{col}', None)
            if pd.isna(val1) or pd.isna(val2):
                continue
            if val1 != val2:
                diffs.append(col)
        return ', '.join(diffs)

    # Добавление столбца с различиями
    df_merged['diff_columns'] = df_merged.apply(get_differences, axis=1)

    return df_merged

resultdf_1 = pd.read_parquet(f"{args.site1}_auxiliary.parquet")
resultdf_2 = pd.read_parquet(f"{args.site2}_auxiliary.parquet")
comp_result = compare_dataframes(resultdf_1, resultdf_2, args.site1, args.site2)

# --------------------------------------------------------------Сохранить результат сравнения в excel--------------------------------------------------------------

def save_comparison_to_excel(df: pd.DataFrame, filename: str):
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    wb = Workbook()
    ws = wb.active

    # Записываем DataFrame в Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    headers = [cell.value for cell in ws[1]]
    diff_col_index = len(headers)
    prefix_columns = [col for col in headers if col != 'Номенклатура' and col != 'diff_columns']

    # --- Автонастройка ширины столбцов по первой строке ---
    for col_idx, cell in enumerate(ws[1], start=1):
        max_length = len(str(cell.value)) if cell.value else 0
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = max_length + 2  # +2 для отступа

    # --- Применение стилей и переносов ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        diff_text = row[diff_col_index - 1].value
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')  # включаем перенос текста

        if isinstance(diff_text, str) and diff_text.strip():
            different_fields = [field.strip() for field in diff_text.split(',')]
            for diff_field in different_fields:
                for col_idx, col_name in enumerate(headers):
                    if col_name.endswith(f"_{diff_field}"):
                        row[col_idx].fill = yellow_fill
        row[diff_col_index - 1].fill = red_fill

    wb.save(filename)

save_comparison_to_excel(comp_result, f'comparison_{args.site1}_vs_{args.site2}.xlsx')